import importlib
import math
import re
import unicodedata
import pandas as pd
import numpy as np
from io import BytesIO
from pathlib import Path
from datetime import datetime, timedelta

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file, current_app
from sqlalchemy import and_, func, select
from . import db
from .models import ConfigList, Colaborador

bp = Blueprint('main', __name__)


last_planilha = None
last_planilha_hc = None


# Helpers


def normalize_matricula(value):
    if value is None:
        return None
    try:
        text = str(value).strip()
        if not text:
            return None
        lowered = text.lower()
        if lowered in {'nan', 'none', 'null'}:
            return None
        numeric = int(float(text))
        if numeric <= 0:
            return None
        return numeric
    except (ValueError, TypeError):
        return None


def normalize_situacao_hc(value):
    """Normaliza os rótulos da coluna "Situação HC" para uso consistente no painel."""
    temporario_label = 'Tempórario'
    if value is None:
        return temporario_label

    try:
        text = str(value).strip()
    except Exception:
        return temporario_label

    if not text:
        return temporario_label

    lowered = text.lower()
    if lowered in {'nan', 'none', 'null'}:
        return temporario_label

    normalized = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
    normalized = normalized.replace('\\', '/').upper()
    normalized = re.sub(r'\s+', ' ', normalized).strip()

    if not normalized:
        return temporario_label

    if normalized in {'N/D', 'ND', 'N A', 'N/A'}:
        return temporario_label

    if normalized in {'SEM INFORMACAO', 'SEM INFORMACOES', 'SEM NADA', 'SEM DADO', 'SEM DADOS', 'SEM REGISTRO'}:
        return temporario_label

    if normalized == 'ATIVIDADE NORMAL':
        return 'Ativo'

    if normalized.startswith('AFASTAMENTO'):
        return 'Afastado'

    if normalized.startswith('FERIAS'):
        return 'Férias'

    if normalized.startswith('RESCISAO'):
        return 'Rescisão'

    return text


def build_execucao_por_voz_lookup(df: pd.DataFrame | None):
    """Constrói uma tabela auxiliar com "Execução por Voz" indexada por Matrícula."""
    if df is None:
        return None

    required_columns = {'Funcionário', 'Execução por Voz'}
    if not required_columns.issubset(df.columns):
        return None

    lookup = df[['Funcionário', 'Execução por Voz']].copy()
    lookup['Funcionário'] = lookup['Funcionário'].apply(normalize_matricula)
    lookup = lookup[lookup['Funcionário'].notna()].copy()
    if lookup.empty:
        return None

    try:
        lookup['Funcionário'] = lookup['Funcionário'].astype(int)
    except Exception:
        pass

    def normalize_execucao(value):
        if pd.isna(value):
            return ''
        text = str(value).strip()
        lowered = text.lower()
        if lowered in {'', 'nan', 'none', 'null'}:
            return ''
        return text

    lookup['Execução por Voz'] = lookup['Execução por Voz'].apply(normalize_execucao)
    lookup['__priority'] = lookup['Execução por Voz'].eq('').astype(int)
    lookup = (
        lookup
        .sort_values(['Funcionário', '__priority', 'Execução por Voz'])
        .drop_duplicates(subset=['Funcionário'], keep='first')
        .drop(columns='__priority')
    )

    return lookup.rename(columns={'Funcionário': 'Matrícula'})


def get_input_column_definitions():
    return [
        {"name": "Do Endereço", "param": "do_endereco", "icon": "geo-alt", "placeholder": "Endereço"},
        {"name": "Funcionário", "param": "funcionario", "icon": "hash", "placeholder": "Funcionário"},
        {"name": "Nome", "param": "nome", "icon": "person", "placeholder": "Nome"},
        {"name": "Data", "param": "data", "icon": "calendar-event", "placeholder": "Data"},
        {"name": "Execução por Voz", "param": "execucao", "icon": "mic", "placeholder": "Execução"},
        {"name": "Treinado", "param": "treinado", "icon": "mortarboard", "placeholder": "Treinado"},
        {"name": "Turno HC", "param": "turno_hc", "icon": "clock-history", "placeholder": "Turno"},
    ]


def slugify_column(label):
    if label is None:
        return ''
    text = str(label)
    normalized = unicodedata.normalize('NFKD', text)
    ascii_text = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    slug = re.sub(r'[^a-z0-9]+', '_', ascii_text.lower()).strip('_')
    return slug or 'col'


def sort_dataframe(df, column, ascending=True):
    if column not in df.columns or df.empty:
        return df
    series = df[column]
    numeric_series = pd.to_numeric(series, errors='coerce')
    if numeric_series.notna().any():
        sort_key = numeric_series
    else:
        datetime_series = pd.to_datetime(series, errors='coerce')
        if datetime_series.notna().any():
            sort_key = datetime_series
        else:
            sort_key = series.astype(str).str.lower()
    sorted_df = df.assign(__sort_key=sort_key)
    sorted_df = sorted_df.sort_values('__sort_key', ascending=ascending, na_position='last', kind='mergesort')
    return sorted_df.drop(columns='__sort_key')


def dataframe_to_excel_response(df: pd.DataFrame, *, filename_prefix: str, sheet_name: str):
    if df is None:
        df = pd.DataFrame()
    df = df.copy()
    try:
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError('Dependência openpyxl não encontrada. Instale com: pip install openpyxl') from exc

    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        header_font = Font(bold=True, color='FFFFFFFF')
        header_fill = PatternFill(start_color='FF0D6EFD', end_color='FF0D6EFD', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        for cell in next(worksheet.iter_rows(min_row=1, max_row=1)):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        worksheet.freeze_panes = 'A2'

        for idx, column in enumerate(df.columns, start=1):
            try:
                series = df[column].astype(str)
                max_length = series.map(len).max()
            except Exception:
                max_length = None
            header_length = len(str(column))
            if max_length is None or pd.isna(max_length):
                max_length = 0
            width = min(max(header_length, max_length) + 2, 60)
            worksheet.column_dimensions[get_column_letter(idx)].width = width

        worksheet.auto_filter.ref = worksheet.dimensions

    buffer.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"{filename_prefix}_{timestamp}.xlsx"
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def manipular_dados(df):
    """Prepara o DF da planilha e faz merge com o banco (apenas tipo TALKMAN) por Matrícula.

    - Gera coluna "Matrícula" a partir de "Funcionário" se necessário.
    - Deduplica por "Matrícula" na planilha.
    - Busca do SQLite (tabela Colaborador) apenas registros com tipo == 'TALKMAN'.
    - Faz merge inner em "Matrícula" e retorna o DF resultante.
    """
    import pandas as pd
    from sqlalchemy import select

    # 1) Preparar DataFrame da planilha
    try:
        database = df.copy()


        database["Funcionário"] = database["Funcionário"].astype(int)
        # Deduplicar por matrícula
        database = database.drop_duplicates(subset=["Funcionário"]).reset_index(drop=True)
    except Exception as e:
        current_app.logger.exception("Falha preparando DataFrame para merge: %s", e)
        flash(f'Falha preparando planilha para merge: {e}', 'danger')
        return None

    try:
        bind = db.session.get_bind()
        stmt = select(
            Colaborador.matricula.label("Matrícula"),
            Colaborador.nome.label("Nome_DB"),
            Colaborador.tipo.label("Tipo"),
            Colaborador.setor.label("Setor"),
            Colaborador.area.label("Área"),
            Colaborador.turno.label("Turno"),
            Colaborador.supervisor.label("Supervisor"),
            Colaborador.integracao.label("Integração"),
            Colaborador.data.label("Data_DB"),
        ).where(Colaborador.tipo == 'TALKMAN')
        df_db = pd.read_sql(stmt, bind)
    except Exception as e:
        current_app.logger.exception("Falha ao ler dados do banco para merge: %s", e)
        flash(f'Falha ao carregar dados do banco: {e}', 'danger')
        return None

    try:
        merged = pd.merge(database, df_db, left_on="Funcionário", right_on="Matrícula", how="left")


    except Exception as e:
        current_app.logger.exception("Falha no merge dos dados: %s", e)
        flash(f'Falha ao mesclar dados: {e}', 'danger')
        return None

    current_app.logger.info("Merge TALKMAN concluído: %s linhas x %s colunas", merged.shape[0], merged.shape[1])
    return merged, database, df_db



def get_list(nome: str) -> list[str]:
    rows = ConfigList.query.filter_by(nome_lista=nome).order_by(ConfigList.valor.asc()).all()
    return [r.valor for r in rows]


@bp.route('/')
def home():
    return redirect(url_for('main.alimentacao'))


@bp.route('/alimentacao', methods=['GET', 'POST'])
def alimentacao():
    lists = {
        'tipo': get_list('tipo'),
        'setor': get_list('setor'),
        'area': get_list('area'),
        'turno': get_list('turno'),
        'integracao': get_list('integracao'),
    }

    if request.method == 'POST':
        try:
            matricula = int(request.form.get('matricula', '').strip())
        except ValueError:
            flash('Matrícula inválida.', 'danger')
            return render_template('alimentacao.html', lists=lists)

        nome = request.form.get('nome', '').strip()
        tipo = request.form.get('tipo', '')
        setor = request.form.get('setor', '')
        area = request.form.get('area', '')
        turno = request.form.get('turno', '')
        supervisor = request.form.get('supervisor', '').strip().upper()
        integracao = request.form.get('integracao', '')
        data_str = request.form.get('data', '')
        observacao = request.form.get('observacao', '').strip()

        # Basic validations
        errors = []
        if not nome:
            errors.append('Nome é obrigatório.')
        if tipo not in lists['tipo']:
            errors.append('Tipo inválido.')
        if setor not in lists['setor']:
            errors.append('Setor inválido.')
        if area not in lists['area']:
            errors.append('Área inválida.')
        if turno not in lists['turno']:
            errors.append('Turno inválido.')
        if integracao not in lists['integracao']:
            errors.append('Integração inválida.')

        try:
            data = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            errors.append('Data inválida.')
            data = None

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('alimentacao.html', lists=lists, form=request.form)

        col = Colaborador(
            matricula=matricula,
            nome=nome,
            tipo=tipo,
            setor=setor,
            area=area,
            turno=turno,
            supervisor=supervisor,
            integracao=integracao,
            data=data,
            observacao=observacao or None,
        )
        db.session.add(col)
        db.session.commit()
        flash('Registro salvo com sucesso.', 'success')
        return redirect(url_for('main.alimentacao'))

    return render_template('alimentacao.html', lists=lists)


@bp.route('/tabela')
def tabela():
    # Período padrão: últimos 30 dias incluindo hoje
    today = datetime.today().date()
    default_min = (today - timedelta(days=29)).strftime('%Y-%m-%d')
    default_max = today.strftime('%Y-%m-%d')
    min_date_str = request.args.get('min_data') or default_min
    max_date_str = request.args.get('max_data') or default_max

    q_nome = (request.args.get('q_nome') or '').strip()
    q_supervisor = (request.args.get('q_supervisor') or '').strip()
    q_matricula_raw = (request.args.get('q_matricula') or '').strip()

    q = Colaborador.query
    if min_date_str:
        try:
            min_date = datetime.strptime(min_date_str, '%Y-%m-%d').date()
            q = q.filter(Colaborador.data >= min_date)
        except ValueError:
            flash('Data mínima inválida.', 'warning')
    if max_date_str:
        try:
            max_date = datetime.strptime(max_date_str, '%Y-%m-%d').date()
            q = q.filter(Colaborador.data <= max_date)
        except ValueError:
            flash('Data máxima inválida.', 'warning')

    # Filtros de texto
    if q_nome:
        q = q.filter(Colaborador.nome.ilike(f"%{q_nome}%"))
    if q_supervisor:
        q = q.filter(Colaborador.supervisor.ilike(f"%{q_supervisor}%"))
    if q_matricula_raw:
        try:
            q_matricula = int(q_matricula_raw)
            q = q.filter(Colaborador.matricula == q_matricula)
        except ValueError:
            flash('Matrícula para filtro deve ser numérica.', 'warning')

    # Paginação
    page = max(1, request.args.get('page', default=1, type=int) or 1)
    # Limita per_page entre 5 e 100
    per_page = request.args.get('per_page', default=25, type=int) or 25
    per_page = max(5, min(per_page, 100))
    q = q.order_by(Colaborador.data.desc(), Colaborador.created_at.desc())
    pagination = db.paginate(q, page=page, per_page=per_page, error_out=False)

    # Janela de páginas para paginação (evita usar max/min em Jinja)
    try:
        total_pages = int(pagination.pages or 1)
    except Exception:
        total_pages = 1
    window = 2
    start_page = max(1, page - window)
    end_page = min(total_pages, page + window)

    return render_template(
        'tabela.html',
        rows=pagination.items,
        min_data=min_date_str,
        max_data=max_date_str,
        q_nome=q_nome,
        q_supervisor=q_supervisor,
        q_matricula=q_matricula_raw,
        pagination=pagination,
        page=page,
        per_page=per_page,
        start_page=start_page,
        end_page=end_page,
    )


@bp.route('/tabela/<int:item_id>/editar', methods=['GET', 'POST'])
def editar_colaborador(item_id: int):
    col = Colaborador.query.get_or_404(item_id)

    lists = {
        'tipo': get_list('tipo'),
        'setor': get_list('setor'),
        'area': get_list('area'),
        'turno': get_list('turno'),
        'integracao': get_list('integracao'),
    }

    for key in ('tipo', 'setor', 'area', 'turno', 'integracao'):
        current_value = getattr(col, key, None)
        if current_value and current_value not in lists[key]:
            lists[key] = [current_value] + lists[key]

    filters_raw = {
        'min_data': request.args.get('min_data'),
        'max_data': request.args.get('max_data'),
        'q_nome': request.args.get('q_nome'),
        'q_matricula': request.args.get('q_matricula'),
        'q_supervisor': request.args.get('q_supervisor'),
        'page': request.args.get('page', type=int),
        'per_page': request.args.get('per_page', type=int),
    }
    filters = {k: v for k, v in filters_raw.items() if v not in (None, '')}
    return_url = url_for('main.tabela', **filters) if filters else url_for('main.tabela')
    form_action = url_for('main.editar_colaborador', item_id=col.id, **filters) if filters else url_for('main.editar_colaborador', item_id=col.id)

    if request.method == 'POST':
        errors = []
        try:
            matricula_raw = request.form.get('matricula', '').strip()
            matricula = int(matricula_raw)
        except ValueError:
            errors.append('Matrícula inválida.')
            matricula = None

        nome = request.form.get('nome', '').strip()
        if not nome:
            errors.append('Nome é obrigatório.')

        tipo = request.form.get('tipo', '')
        setor = request.form.get('setor', '')
        area = request.form.get('area', '')
        turno = request.form.get('turno', '')
        supervisor = request.form.get('supervisor', '').strip().upper()
        integracao = request.form.get('integracao', '')
        data_str = request.form.get('data', '')
        observacao = request.form.get('observacao', '').strip()

        if tipo not in lists['tipo']:
            errors.append('Tipo inválido.')
        if setor not in lists['setor']:
            errors.append('Setor inválido.')
        if area not in lists['area']:
            errors.append('Área inválida.')
        if turno not in lists['turno']:
            errors.append('Turno inválido.')
        if integracao not in lists['integracao']:
            errors.append('Integração inválida.')

        try:
            data = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            errors.append('Data inválida.')
            data = None

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template(
                'editar_colaborador.html',
                col=col,
                lists=lists,
                form=request.form,
                form_action=form_action,
                return_url=return_url,
            )

        if matricula is not None:
            col.matricula = matricula
        col.nome = nome
        col.tipo = tipo
        col.setor = setor
        col.area = area
        col.turno = turno
        col.supervisor = supervisor
        col.integracao = integracao
        if data:
            col.data = data
        col.observacao = observacao or None

        db.session.commit()
        flash('Registro atualizado com sucesso.', 'success')
        return redirect(return_url)

    form_data = {
        'matricula': col.matricula,
        'nome': col.nome,
        'tipo': col.tipo,
        'setor': col.setor,
        'area': col.area,
        'turno': col.turno,
        'supervisor': col.supervisor,
        'integracao': col.integracao,
        'data': col.data.strftime('%Y-%m-%d') if col.data else '',
        'observacao': col.observacao or '',
    }

    return render_template(
        'editar_colaborador.html',
        col=col,
        lists=lists,
        form=form_data,
        form_action=form_action,
        return_url=return_url,
    )





@bp.route('/excluir/<int:item_id>', methods=['POST'])
def excluir(item_id: int):
    item = Colaborador.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash('Registro excluído com sucesso.', 'success')

    # Preserva filtros/paginação vindos por query string
    min_data = request.args.get('min_data')
    max_data = request.args.get('max_data')
    q_nome = request.args.get('q_nome')
    q_matricula = request.args.get('q_matricula')
    q_supervisor = request.args.get('q_supervisor')
    page = request.args.get('page', type=int)
    per_page = request.args.get('per_page', type=int)
    args = {k: v for k, v in {
        'min_data': min_data,
        'max_data': max_data,
        'q_nome': q_nome,
        'q_matricula': q_matricula,
        'q_supervisor': q_supervisor,
        'page': page,
        'per_page': per_page,
    }.items() if v}
    return redirect(url_for('main.tabela', **args))


@bp.route('/tabela/export')
def tabela_export():
    """Exporta os dados filtrados para XLSX."""
    from io import BytesIO
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except Exception:
        return jsonify({
            'error': 'Dependência openpyxl não encontrada. Instale com: pip install openpyxl'
        }), 500

    # Período padrão: últimos 30 dias incluindo hoje
    today = datetime.today().date()
    default_min = (today - timedelta(days=29)).strftime('%Y-%m-%d')
    default_max = today.strftime('%Y-%m-%d')
    min_date_str = request.args.get('min_data') or default_min
    max_date_str = request.args.get('max_data') or default_max

    q_nome = (request.args.get('q_nome') or '').strip()
    q_supervisor = (request.args.get('q_supervisor') or '').strip()
    q_matricula_raw = (request.args.get('q_matricula') or '').strip()

    q = Colaborador.query
    if min_date_str:
        try:
            min_date = datetime.strptime(min_date_str, '%Y-%m-%d').date()
            q = q.filter(Colaborador.data >= min_date)
        except ValueError:
            pass
    if max_date_str:
        try:
            max_date = datetime.strptime(max_date_str, '%Y-%m-%d').date()
            q = q.filter(Colaborador.data <= max_date)
        except ValueError:
            pass

    if q_nome:
        q = q.filter(Colaborador.nome.ilike(f"%{q_nome}%"))
    if q_supervisor:
        q = q.filter(Colaborador.supervisor.ilike(f"%{q_supervisor}%"))
    if q_matricula_raw:
        try:
            q_matricula = int(q_matricula_raw)
            q = q.filter(Colaborador.matricula == q_matricula)
        except ValueError:
            pass

    q = q.order_by(Colaborador.data.desc(), Colaborador.created_at.desc())
    rows = q.all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dados'
    headers = [
        'Data', 'Matrícula', 'Nome', 'Tipo', 'Setor', 'Área', 'Turno', 'Supervisor', 'Integração', 'Observação'
    ]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.data.strftime('%Y-%m-%d') if r.data else '',
            r.matricula,
            r.nome,
            r.tipo,
            r.setor,
            r.area,
            r.turno,
            r.supervisor,
            r.integracao,
            r.observacao or '',
        ])
    # Largura das colunas básica
    widths = [12, 10, 26, 14, 18, 18, 12, 18, 12, 40]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = 'tabela_qualidade.xlsx'
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# Config Lists API and page
@bp.route('/config-lists')
def config_lists():
    # Show grouped lists
    all_lists = {}
    for nome in ['tipo', 'setor', 'area', 'turno', 'integracao']:
        all_lists[nome] = get_list(nome)
    return render_template('config_lists.html', all_lists=all_lists)


@bp.route('/api/lists/<nome_lista>', methods=['GET', 'POST', 'PUT', 'DELETE'])
def api_lists(nome_lista):
    nome_lista = nome_lista.lower()
    if request.method == 'GET':
        return jsonify(get_list(nome_lista))

    data = request.get_json(force=True, silent=True) or {}

    if request.method == 'POST':
        valor = (data.get('valor') or '').strip()
        if not valor:
            return jsonify({'error': 'Valor é obrigatório'}), 400
        if len(valor) > 120:
            return jsonify({'error': 'Valor excede 120 caracteres'}), 400
        # Verificação case-insensitive
        exists = (
            db.session.query(ConfigList.id)
            .filter(
                ConfigList.nome_lista == nome_lista,
                func.lower(ConfigList.valor) == func.lower(valor),
            )
            .first()
        )
        if exists:
            return jsonify({'error': 'Valor já existe (comparação sem diferenciar maiúsculas/minúsculas)'}), 409
        db.session.add(ConfigList(nome_lista=nome_lista, valor=valor))
        db.session.commit()
        return jsonify({'ok': True})

    if request.method == 'PUT':
        old = (data.get('old') or '').strip()
        new = (data.get('new') or '').strip()
        if not old or not new:
            return jsonify({'error': 'Parâmetros inválidos'}), 400
        if len(new) > 120:
            return jsonify({'error': 'Valor excede 120 caracteres'}), 400
        row = ConfigList.query.filter_by(nome_lista=nome_lista, valor=old).first()
        if not row:
            return jsonify({'error': 'Valor não encontrado'}), 404
        # Case-insensitive uniqueness
        exists = (
            db.session.query(ConfigList.id)
            .filter(
                ConfigList.nome_lista == nome_lista,
                func.lower(ConfigList.valor) == func.lower(new),
                ConfigList.id != row.id,
            )
            .first()
        )
        if exists:
            return jsonify({'error': 'Novo valor já existe (comparação sem diferenciar maiúsculas/minúsculas)'}), 409
        row.valor = new
        db.session.commit()
        return jsonify({'ok': True})

    if request.method == 'DELETE':
        valor = (data.get('valor') or '').strip()
        row = ConfigList.query.filter_by(nome_lista=nome_lista, valor=valor).first()
        if not row:
            return jsonify({'error': 'Valor não encontrado'}), 404
        # Bloqueia remoção se valor estiver em uso por Colaborador
        field_map = {
            'tipo': Colaborador.tipo,
            'setor': Colaborador.setor,
            'area': Colaborador.area,
            'turno': Colaborador.turno,
            'integracao': Colaborador.integracao,
        }
        if nome_lista in field_map:
            in_use = db.session.query(Colaborador.id).filter(field_map[nome_lista] == valor).first()
            if in_use:
                return jsonify({'error': 'Não é possível remover: valor está em uso em registros existentes'}), 409
        db.session.delete(row)
        db.session.commit()
        return jsonify({'ok': True})

    return jsonify({'error': 'Método não suportado'}), 405


# Página para upload de planilha (Input*Dados)
@bp.route('/input-dados', methods=['GET', 'POST'])
def input_dados():
    if request.method == 'POST':
        files = [f for f in request.files.getlist('files') if f and f.filename]
        if not files:
            flash('Nenhum arquivo selecionado.', 'warning')
            return redirect(url_for('main.input_dados'))

        allowed_extensions = {'.xlsx', '.xls', '.xlsb'}

        try:
            import pandas as pd  # import local para não quebrar app se pandas não estiver instalado
        except Exception:
            flash('Dependência pandas não encontrada. Instale com: pip install pandas', 'danger')
            return redirect(url_for('main.input_dados'))

        global last_planilha, last_planilha_hc
        last_planilha_hc = None

        def determine_engine(ext: str) -> str:
            ext = (ext or '').lower()
            if ext == '.xlsb':
                try:
                    importlib.import_module('pyxlsb')
                except ImportError:
                    raise RuntimeError('Dependência pyxlsb não encontrada. Instale com: pip install pyxlsb')
                return 'pyxlsb'
            if ext == '.xls':
                try:
                    importlib.import_module('xlrd')
                except ImportError:
                    raise RuntimeError('Dependência xlrd não encontrada. Instale com: pip install xlrd==1.2.0')
                return 'xlrd'
            try:
                importlib.import_module('openpyxl')
            except ImportError:
                raise RuntimeError('Dependência openpyxl não encontrada. Instale com: pip install openpyxl')
            return 'openpyxl'

        def read_dataframe(storage, *, sheet_name=0, extension=None):
            ext = extension or Path(storage.filename).suffix.lower()
            engine = determine_engine(ext)
            try:
                storage.stream.seek(0)
            except Exception:
                pass
            return pd.read_excel(storage, engine=engine, sheet_name=sheet_name)

        preview_filename = None
        preview_df = None
        preview_shape = None
        processed_any = False
        invalid_names = []
        hc_previews = []

        for file in files:
            filename = file.filename
            extension = Path(filename).suffix.lower()
            if extension not in allowed_extensions:
                invalid_names.append(filename)
                continue

            uppercase_name = filename.strip().upper()

            if uppercase_name.startswith('HC'):
                try:
                    df_hc = read_dataframe(file, sheet_name='Base Colab.', extension=extension)
                except RuntimeError as dep_err:
                    flash(str(dep_err), 'danger')
                    continue
                except ValueError as sheet_err:
                    flash(f'Planilha "{filename}" não contém a aba "Base Colab.": {sheet_err}', 'danger')
                    continue
                except Exception as err:
                    current_app.logger.exception('Falha ao carregar planilha HC %s', filename)
                    flash(f'Falha ao processar a planilha "{filename}": {err}', 'danger')
                    continue

                display_df = df_hc.copy()
                expected_cols = ["Matrícula", "Cargo", "Situação", "Turno"]
                missing_cols = [col for col in expected_cols if col not in display_df.columns]
                if missing_cols:
                    flash(f'Planilha "{filename}" não possui as colunas esperadas: {", ".join(missing_cols)}', 'warning')
                    continue

                display_df = display_df[expected_cols].copy()
                display_df = display_df.rename(columns={
                    "Cargo": "Cargo HC",
                    "Situação": "Situação HC",
                    "Turno": "Turno HC"
                })
                display_df['Situação HC'] = display_df['Situação HC'].apply(normalize_situacao_hc)
                display_df['Matrícula'] = display_df['Matrícula'].apply(normalize_matricula)
                display_df = display_df[display_df['Matrícula'].notna()].copy()
                try:
                    display_df['Matrícula'] = display_df['Matrícula'].astype(int)
                except Exception:
                    pass

                processed_any = True
                last_planilha_hc = display_df.copy()
                current_app.logger.info('Planilha HC detectada: "%s" (%s). Linhas: %s | Colunas: %s', filename, extension or 'sem extensão', display_df.shape[0], list(display_df.columns))
                try:
                    preview_block = display_df.head(5).copy()
                except Exception:
                    preview_block = display_df
                try:
                    preview_block = preview_block.fillna('')
                except Exception:
                    pass
                try:
                    preview_rows = preview_block.astype(str).values.tolist()
                except Exception:
                    preview_rows = preview_block.values.tolist()
                preview_cols_hc = [str(c) for c in list(preview_block.columns)]
                hc_previews.append({
                    'filename': filename,
                    'shape': display_df.shape,
                    'columns': preview_cols_hc,
                    'rows': preview_rows,
                })
                try:
                    console_preview = display_df.head(5).to_string(index=False)
                except Exception:
                    console_preview = str(display_df.head(5))
                print(f'[Input*Dados][HC] {filename} - Prévia das 5 primeiras linhas:\n{console_preview}')
                flash(f'Planilha "{filename}" (HC) carregada e registrada no console.', 'info')
                continue

            try:
                df = read_dataframe(file, extension=extension)
            except RuntimeError as dep_err:
                flash(str(dep_err), 'danger')
                continue
            except Exception as err:
                current_app.logger.exception('Falha ao processar planilha %s', filename)
                flash(f'Falha ao processar o arquivo "{filename}": {err}', 'danger')
                continue

            processed_any = True

            if filename.startswith("Rastreabilidade_Tra"):
                try:
                    df_listColomns = ["Do Endereço", "Funcionário", "Nome", "Data", "Execução por Voz"]
                    df_trabalho = df[df_listColomns].copy()
                    df_trabalho["MOD"] = df_trabalho["Do Endereço"].fillna("").astype(str).str[:1]

                    existing_matriculas = set()
                    for (value,) in (
                        db.session.query(Colaborador.matricula)
                        .filter(Colaborador.matricula.isnot(None))
                        .all()
                    ):
                        if value is None:
                            continue
                        try:
                            existing_matriculas.add(int(value))
                        except (TypeError, ValueError):
                            continue

                    def flag_treinado(raw):
                        matricula = normalize_matricula(raw)
                        return 'Sim' if matricula is not None and matricula in existing_matriculas else 'Não'

                    df_trabalho['Treinado'] = df_trabalho['Funcionário'].apply(flag_treinado)

                    flash(f'Arquivo de rastreabilidade detectado. Linhas: Columns {df_listColomns} | MOD e Treinado adicionados', 'info')

                    global last_planilha
                    planilha = df_trabalho.copy()
                    resultado = manipular_dados(planilha)
                    if resultado is not None:
                        df_manipulada, planilha, bancodb = resultado
                        if planilha is not None:
                            last_planilha = planilha
                except Exception as e:
                    current_app.logger.exception('Falha ao processar arquivo de rastreabilidade %s', filename)
                    flash(f'Falha ao processar arquivo de rastreabilidade "{filename}": {e}', 'danger')
                    continue

                candidate_df = df_trabalho
            else:
                candidate_df = df

            if preview_df is None:
                preview_df = candidate_df
                preview_filename = filename
                preview_shape = candidate_df.shape

            rows_count = candidate_df.shape[0]
            cols_count = candidate_df.shape[1]
            flash(f'Arquivo "{filename}" processado com sucesso. Linhas: {rows_count} | Colunas: {cols_count}', 'success')

        if invalid_names:
            ignored = ', '.join(invalid_names)
            flash(f'Arquivos ignorados por formato inválido: {ignored}', 'warning')

        if not processed_any:
            return redirect(url_for('main.input_dados'))

        if preview_df is None and not hc_previews:
            return redirect(url_for('main.input_dados'))

        preview_cols = None
        preview_rows = None
        if preview_df is not None:
            preview_display = preview_df.head(5).copy()
            try:
                preview_display = preview_display.fillna('')
            except Exception:
                pass
            try:
                preview_rows = preview_display.astype(str).values.tolist()
            except Exception:
                preview_rows = preview_display.values.tolist()
            preview_cols = [str(c) for c in list(preview_display.columns)]

        return render_template(
            'input_dados.html',
            preview_cols=preview_cols,
            preview_rows=preview_rows,
            preview_shape=preview_shape,
            filename=preview_filename,
            hc_previews=hc_previews,
        )

    return render_template('input_dados.html')

@bp.route('/painel-grafico', methods=['GET'])
def painel_grafico(planilha=None):
    """Renderiza o painel gráfico com cards de consolidados."""

    # Consolidados do banco com período selecionável
    try:
        # Faixa total disponível no banco (para exibição informativa)
        min_all = db.session.query(func.min(Colaborador.data)).scalar()
        max_all = db.session.query(func.max(Colaborador.data)).scalar()

        # Turnos disponíveis (distintos no banco)
        turnos_rows = (
            db.session.query(Colaborador.turno)
            .filter(Colaborador.turno.isnot(None))
            .distinct()
            .order_by(Colaborador.turno.asc())
            .all()
        )
        available_turnos = [t[0] for t in turnos_rows]

        # Filtros adicionais: Setor, Tipo, Supervisor (para timeline)
        setores_rows = (
            db.session.query(Colaborador.setor)
            .filter(Colaborador.setor.isnot(None))
            .distinct()
            .order_by(Colaborador.setor.asc())
            .all()
        )
        tipos_rows = (
            db.session.query(Colaborador.tipo)
            .filter(Colaborador.tipo.isnot(None))
            .distinct()
            .order_by(Colaborador.tipo.asc())
            .all()
        )
        supervisores_rows = (
            db.session.query(Colaborador.supervisor)
            .filter(Colaborador.supervisor.isnot(None))
            .distinct()
            .order_by(Colaborador.supervisor.asc())
            .all()
        )
        available_setores = [s[0] for s in setores_rows]
        available_tipos = [t[0] for t in tipos_rows]
        available_supervisores = [s[0] for s in supervisores_rows]

        # Ler período do usuário (GET) e definir padrão como HOJE (performance)
        min_param = request.args.get('min_data')
        max_param = request.args.get('max_data')
        selected_turno = request.args.get('turno') or 'all'
        selected_setor = request.args.get('setor') or 'all'
        selected_tipo = request.args.get('tipo') or 'all'
        selected_supervisor = request.args.get('supervisor') or 'all'

        def parse_date(s):
            if not s:
                return None
            try:
                return datetime.strptime(s, '%Y-%m-%d').date()
            except Exception:
                return None

        today = datetime.today().date()
        sel_min = parse_date(min_param) or today
        sel_max = parse_date(max_param) or today

        # Query filtrada pelo período e turno selecionados
        q = db.session.query(Colaborador)
        if sel_min:
            q = q.filter(Colaborador.data >= sel_min)
        if sel_max:
            q = q.filter(Colaborador.data <= sel_max)
        if selected_turno and selected_turno != 'all':
            q = q.filter(Colaborador.turno == selected_turno)
        if selected_tipo and selected_tipo != 'all':
            q = q.filter(Colaborador.tipo == selected_tipo)

        total_colaboradores = q.count()

        # Agregações para gráficos (aplicando os mesmos filtros)
        q_setor = db.session.query(
            Colaborador.setor.label('setor'),
            func.count(func.distinct(Colaborador.matricula)).label('qtd')
        )
        if sel_min:
            q_setor = q_setor.filter(Colaborador.data >= sel_min)
        if sel_max:
            q_setor = q_setor.filter(Colaborador.data <= sel_max)
        if selected_turno and selected_turno != 'all':
            q_setor = q_setor.filter(Colaborador.turno == selected_turno)
        if selected_tipo and selected_tipo != 'all':
            q_setor = q_setor.filter(Colaborador.tipo == selected_tipo)
        q_setor = q_setor.filter(Colaborador.setor.isnot(None)).group_by(Colaborador.setor).order_by(func.count(func.distinct(Colaborador.matricula)).desc())
        setor_rows = q_setor.all()
        setor_labels = [r.setor for r in setor_rows]
        setor_series = [int(r.qtd or 0) for r in setor_rows]

        # Agregação simples por tipo (contagem total sem distinct)
        q_tipo_resumo = db.session.query(
            Colaborador.tipo.label('tipo'),
            func.count(Colaborador.matricula).label('qtd')
        )
        if sel_min:
            q_tipo_resumo = q_tipo_resumo.filter(Colaborador.data >= sel_min)
        if sel_max:
            q_tipo_resumo = q_tipo_resumo.filter(Colaborador.data <= sel_max)
        if selected_turno and selected_turno != 'all':
            q_tipo_resumo = q_tipo_resumo.filter(Colaborador.turno == selected_turno)
        if selected_tipo and selected_tipo != 'all':
            q_tipo_resumo = q_tipo_resumo.filter(Colaborador.tipo == selected_tipo)
        q_tipo_resumo = (
            q_tipo_resumo
            .filter(Colaborador.tipo.isnot(None))
            .group_by(Colaborador.tipo)
            .order_by(func.count(Colaborador.matricula).desc())
        )
        tipo_rows = q_tipo_resumo.all()
        tipo_labels = [r.tipo for r in tipo_rows]
        tipo_series = [int(r.qtd or 0) for r in tipo_rows]

        q_turno = db.session.query(
            Colaborador.turno.label('turno'),
            func.count(Colaborador.matricula).label('qtd')
        )
        if sel_min:
            q_turno = q_turno.filter(Colaborador.data >= sel_min)
        if sel_max:
            q_turno = q_turno.filter(Colaborador.data <= sel_max)
        if selected_turno and selected_turno != 'all':
            q_turno = q_turno.filter(Colaborador.turno == selected_turno)
        if selected_tipo and selected_tipo != 'all':
            q_turno = q_turno.filter(Colaborador.tipo == selected_tipo)
        q_turno = q_turno.filter(Colaborador.turno.isnot(None)).group_by(Colaborador.turno).order_by(func.count(Colaborador.matricula).desc())
        turno_rows = q_turno.all()
        turno_labels = [r.turno for r in turno_rows]
        turno_series = [int(r.qtd or 0) for r in turno_rows]

        # Agregação simples por setor (contagem total sem distinct para análise de volume)
        q_setor_total = db.session.query(
            Colaborador.setor.label('setor'),
            func.count(Colaborador.id).label('qtd')
        )
        if sel_min:
            q_setor_total = q_setor_total.filter(Colaborador.data >= sel_min)
        if sel_max:
            q_setor_total = q_setor_total.filter(Colaborador.data <= sel_max)
        if selected_turno and selected_turno != 'all':
            q_setor_total = q_setor_total.filter(Colaborador.turno == selected_turno)
        if selected_tipo and selected_tipo != 'all':
            q_setor_total = q_setor_total.filter(Colaborador.tipo == selected_tipo)
        q_setor_total = (
            q_setor_total
            .filter(Colaborador.setor.isnot(None))
            .group_by(Colaborador.setor)
            .order_by(func.count(Colaborador.id).desc())
        )
        setor_total_rows = q_setor_total.all()
        stacked_categories = [r.setor for r in setor_total_rows]
        stacked_series = [int(r.qtd or 0) for r in setor_total_rows]
        min_data = db.session.query(func.min(Colaborador.data))
        max_data = db.session.query(func.max(Colaborador.data))
        if sel_min:
            min_data = min_data.filter(Colaborador.data >= sel_min)
            max_data = max_data.filter(Colaborador.data >= sel_min)
        if sel_max:
            min_data = min_data.filter(Colaborador.data <= sel_max)
            max_data = max_data.filter(Colaborador.data <= sel_max)
        if selected_turno and selected_turno != 'all':
            min_data = min_data.filter(Colaborador.turno == selected_turno)
            max_data = max_data.filter(Colaborador.turno == selected_turno)
        if selected_tipo and selected_tipo != 'all':
            min_data = min_data.filter(Colaborador.tipo == selected_tipo)
            max_data = max_data.filter(Colaborador.tipo == selected_tipo)
        min_data = min_data.scalar()
        max_data = max_data.scalar()

        # Strings para inputs (YYYY-MM-DD)
        def to_str(d):
            try:
                return d.strftime('%Y-%m-%d') if d else ''
            except Exception:
                return ''

        min_all_str = to_str(min_all)
        max_all_str = to_str(max_all)
        min_data_str = to_str(sel_min)
        max_data_str = to_str(sel_max)
        today_str = to_str(today)

        # Série temporal (Data X contagem distinta de Matrícula) com filtros
        q_time = db.session.query(
            Colaborador.data.label('data'),
            func.count(func.distinct(Colaborador.matricula)).label('qtd'),
        )
        if sel_min:
            q_time = q_time.filter(Colaborador.data >= sel_min)
        if sel_max:
            q_time = q_time.filter(Colaborador.data <= sel_max)
        if selected_turno and selected_turno != 'all':
            q_time = q_time.filter(Colaborador.turno == selected_turno)
        if selected_setor and selected_setor != 'all':
            q_time = q_time.filter(Colaborador.setor == selected_setor)
        if selected_tipo and selected_tipo != 'all':
            q_time = q_time.filter(Colaborador.tipo == selected_tipo)
        if selected_supervisor and selected_supervisor != 'all':
            q_time = q_time.filter(Colaborador.supervisor == selected_supervisor)
        q_time = q_time.group_by(Colaborador.data).order_by(Colaborador.data.asc())
        time_rows = q_time.all()
        # Converter para pares [timestamp_ms, valor]
        timeline_data = []
        for r in time_rows:
            try:
                # Usar meio-dia para evitar DST edge-cases
                ts = datetime(r.data.year, r.data.month, r.data.day, 12, 0, 0)
                timeline_data.append([int(ts.timestamp() * 1000), int(r.qtd or 0)])
            except Exception:
                pass
    except Exception as e:
        current_app.logger.exception('Falha ao calcular consolidados do banco: %s', e)
        total_colaboradores, min_data, max_data = 0, None, None
        min_all_str = max_all_str = min_data_str = max_data_str = today_str = ''
        available_turnos = []
        selected_turno = 'all'
        available_setores = []
        available_tipos = []
        available_supervisores = []
        selected_setor = selected_tipo = selected_supervisor = 'all'
        timeline_data = []
        setor_labels = []
        setor_series = []
        tipo_labels = []
        tipo_series = []
        turno_labels = []
        turno_series = []
        stacked_categories = []
        stacked_series = []

    def build_query_args(skip_keys=None, overrides=None):
        skip = set(skip_keys or [])
        base = {k: v for k, v in request.args.to_dict(flat=True).items() if v not in (None, '') and k not in skip}
        if overrides:
            base.update(overrides)
        return base

    hc_turno_lookup = {}
    try:
        lookup_source_hc = last_planilha_hc
    except NameError:
        lookup_source_hc = None
    if lookup_source_hc is not None:
        try:
            required_cols = {'Matrícula', 'Turno HC'}
            if required_cols.issubset(lookup_source_hc.columns):
                extra_cols = []
                if 'Situação HC' in lookup_source_hc.columns:
                    extra_cols.append('Situação HC')
                if 'Turno' in lookup_source_hc.columns:
                    extra_cols.append('Turno')
                turno_lookup_df = lookup_source_hc[['Matrícula', 'Turno HC', *extra_cols]].copy()
                turno_lookup_df['Matrícula'] = turno_lookup_df['Matrícula'].apply(normalize_matricula)
                turno_lookup_df = turno_lookup_df[turno_lookup_df['Matrícula'].notna()]
                turno_lookup_df['Turno HC'] = turno_lookup_df['Turno HC'].fillna('').astype(str).str.strip()

                if 'Situação HC' in turno_lookup_df.columns:
                    situacao_normalizada = turno_lookup_df['Situação HC'].apply(normalize_situacao_hc).fillna('')
                    situacao_ascii = (
                        situacao_normalizada
                        .astype(str)
                        .apply(lambda value: unicodedata.normalize('NFKD', value).encode('ASCII', 'ignore').decode('ASCII'))
                        .str.lower()
                    )
                    temporario_mask = situacao_ascii.str.contains('tempor', na=False)
                else:
                    temporario_mask = pd.Series(False, index=turno_lookup_df.index)

                if 'Turno' in turno_lookup_df.columns:
                    turno_fallback = turno_lookup_df['Turno'].fillna('').astype(str).str.strip()
                else:
                    turno_fallback = pd.Series('', index=turno_lookup_df.index)

                fallback_mask = (turno_lookup_df['Turno HC'] == '') & temporario_mask & (turno_fallback != '')
                if fallback_mask.any():
                    turno_lookup_df.loc[fallback_mask, 'Turno HC'] = turno_fallback.loc[fallback_mask]

                turno_lookup_df = turno_lookup_df.drop_duplicates(subset=['Matrícula'], keep='first')
                turno_lookup_df['Turno HC'] = turno_lookup_df['Turno HC'].replace('', np.nan).fillna('1° Turno')
                hc_turno_lookup = turno_lookup_df.set_index('Matrícula')['Turno HC'].to_dict()
        except Exception as err:
            current_app.logger.warning('Falha ao construir lookup de Turno HC para Input*Dados: %s', err)
            hc_turno_lookup = {}

    input_column_definitions = get_input_column_definitions()
    input_table_columns = [col["name"] for col in input_column_definitions]
    input_table_rows = []
    input_table_total = 0
    input_table_page_size = 10
    input_table_page = request.args.get('input_page', default=1, type=int) or 1
    input_table_page = max(1, input_table_page)
    input_table_pages = 0
    input_table_has_data = False
    input_table_pagination = None
    input_table_range_start = 0
    input_table_range_end = 0
    merge_colab_percent = None
    merge_turno_charts = {
        'turno1': {
            'labels': [],
            'values': [],
            'datasets': [],
            'totals': {'execucao': 0, 'treinado': 0},
            'series_label': 'Treinados',
            'turno_label': '1° Turno',
            'color': '#f59e0b'
        },
        'turno2': {
            'labels': [],
            'values': [],
            'datasets': [],
            'totals': {'execucao': 0, 'treinado': 0},
            'series_label': 'Treinados',
            'turno_label': '2° Turno',
            'color': '#ef4444'
        }
    }
    input_sort = (request.args.get('input_sort') or '').strip()
    valid_input_sorts = {col['param'] for col in input_column_definitions}
    if input_sort not in valid_input_sorts:
        input_sort = ''
    input_order = (request.args.get('input_order') or 'asc').lower()
    if input_order not in {'asc', 'desc'}:
        input_order = 'asc'
    if not input_sort:
        input_order = 'asc'
    input_filters = {}
    for col in input_column_definitions:
        value = (request.args.get(f"input_filter_{col['param']}") or '').strip()
        col['filter_value'] = value
        if value:
            input_filters[col['name']] = value

    try:
        source_df = last_planilha
    except NameError:
        source_df = None

    if source_df is not None:
        try:
            df_input = source_df.copy()
            missing_cols = [col for col in input_table_columns if col not in df_input.columns]
            if missing_cols:
                current_app.logger.warning('Planilha Input*Dados ajustada por colunas ausentes: %s', missing_cols)
                for col in missing_cols:
                    default_value = 'Não' if col == 'Treinado' else ''
                    df_input[col] = default_value
            df_input = df_input[input_table_columns].copy()
            if hc_turno_lookup and {'Funcionário', 'Turno HC'}.issubset(df_input.columns):
                try:
                    normalized_matriculas = df_input['Funcionário'].apply(normalize_matricula)
                    mapped_turnos = normalized_matriculas.map(hc_turno_lookup).fillna('')
                    existing_turnos = df_input['Turno HC'].fillna('').astype(str)
                    df_input['Turno HC'] = (
                        existing_turnos
                        .where(existing_turnos.str.strip() != '', mapped_turnos)
                        .fillna('')
                    )
                    df_input['Turno HC'] = df_input['Turno HC'].replace('', '1° Turno')
                except Exception as err:
                    current_app.logger.warning('Falha ao combinar Turno HC com Input*Dados: %s', err)
            filtered_df = df_input.copy()
            for col_name, filter_value in input_filters.items():
                filter_series = filtered_df[col_name].astype(str).fillna('')
                filtered_df = filtered_df[filter_series.str.contains(filter_value, case=False, na=False)]

            if input_sort:
                sort_column = next((col['name'] for col in input_column_definitions if col['param'] == input_sort), None)
                if sort_column:
                    ascending = input_order == 'asc'
                    filtered_df = sort_dataframe(filtered_df, sort_column, ascending=ascending)

            display_df = filtered_df.fillna('')
            input_table_total = len(display_df)
            if input_table_total > 0:
                input_table_has_data = True
                input_table_pages = max(1, math.ceil(input_table_total / input_table_page_size))
                if input_table_page > input_table_pages:
                    input_table_page = input_table_pages
                start = (input_table_page - 1) * input_table_page_size
                end = start + input_table_page_size
                page_df = display_df.iloc[start:end]
                input_table_range_start = start + 1
                input_table_range_end = min(end, input_table_total)
                input_table_rows = []
                for row in page_df.itertuples(index=False, name=None):
                    formatted = {}
                    for col, value in zip(input_table_columns, row):
                        cell = value
                        if cell is None:
                            text = ''
                        elif isinstance(cell, (int, float)):
                            if isinstance(cell, float) and math.isnan(cell):
                                text = ''
                            elif isinstance(cell, float) and cell.is_integer():
                                text = str(int(cell))
                            else:
                                text = str(cell)
                        else:
                            text = str(cell)
                        formatted[col] = text
                    input_table_rows.append(formatted)

                negative_pattern = re.compile(r"\b(n[aã]o|pendente|aguard|sem|falta)\b", re.IGNORECASE)

                def is_execucao_sim(value):
                    if value is None:
                        return False
                    try:
                        text = str(value).strip()
                    except Exception:
                        return False
                    if not text:
                        return False
                    lowered = text.lower()
                    if lowered in {"", "nan", "none", "null", "0"}:
                        return False
                    if negative_pattern.search(lowered):
                        return False
                    return 'sim' in lowered

                trained_mask = display_df["Treinado"].astype(str).str.strip().str.lower() == 'sim'
                trained_total = int(trained_mask.sum())

                if 'Execução por Voz' in display_df.columns:
                    exec_sim_mask = display_df['Execução por Voz'].apply(is_execucao_sim)
                else:
                    exec_sim_mask = pd.Series(False, index=display_df.index)

                trained_exec_mask = trained_mask & exec_sim_mask
                trained_exec_count = int(trained_exec_mask.sum())
                trained_no_exec_count = max(0, trained_total - trained_exec_count)
                total_count = trained_exec_count + trained_no_exec_count
                if total_count > 0:
                    trained_pct = round((trained_exec_count / total_count) * 100, 2)
                    untrained_pct = round((trained_no_exec_count / total_count) * 100, 2)
                else:
                    trained_pct = untrained_pct = 0.0

                merge_colab_percent = {
                    "labels": ["Com execução por Voz", "Sem execução por Voz"],
                    "values": [trained_exec_count, trained_no_exec_count],
                    "percentages": [trained_pct, untrained_pct],
                    "total": total_count
                }

                if {'Data', 'Turno HC'}.issubset(display_df.columns):
                    def normalize_turno_label(value):
                        default_turno = '1° Turno'
                        if value is None:
                            return default_turno
                        try:
                            text = str(value).strip()
                        except Exception:
                            return default_turno
                        if not text:
                            return default_turno
                        normalized = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
                        lowered = normalized.lower().replace('º', '').replace('°', '')
                        lowered = re.sub(r'[^a-z0-9]+', ' ', lowered).strip()
                        if lowered.startswith('1') or lowered.startswith('primeiro') or lowered.startswith('turno 1'):
                            return '1° Turno'
                        if lowered.startswith('2') or lowered.startswith('segundo') or lowered.startswith('turno 2'):
                            return '2° Turno'
                        return text or default_turno

                    def prepare_group(mask: pd.Series, count_label: str) -> pd.DataFrame:
                        if mask is None or not mask.any():
                            return pd.DataFrame(columns=['__turno', '__parsed_date', count_label])
                        subset = display_df.loc[mask, ['Data', 'Turno HC']].copy()
                        subset['__parsed_date'] = pd.to_datetime(subset['Data'], dayfirst=True, errors='coerce')
                        subset = subset.dropna(subset=['__parsed_date'])
                        if subset.empty:
                            return pd.DataFrame(columns=['__turno', '__parsed_date', count_label])
                        subset['__turno'] = subset['Turno HC'].apply(normalize_turno_label)
                        grouped = (
                            subset
                            .groupby(['__turno', '__parsed_date'])
                            .size()
                            .reset_index(name=count_label)
                        )
                        return grouped

                    exec_grouped = prepare_group(exec_sim_mask, 'execucao_count')
                    trained_grouped = prepare_group(trained_mask, 'treinado_count')

                    if not exec_grouped.empty or not trained_grouped.empty:
                        combined = pd.merge(
                            exec_grouped,
                            trained_grouped,
                            on=['__turno', '__parsed_date'],
                            how='outer'
                        ).fillna(0)
                        combined['execucao_count'] = combined.get('execucao_count', 0).astype(int)
                        combined['treinado_count'] = combined.get('treinado_count', 0).astype(int)

                        for key, turno_label in [('turno1', '1° Turno'), ('turno2', '2° Turno')]:
                            turno_df = combined[combined['__turno'] == turno_label].copy()
                            if not turno_df.empty:
                                turno_df = turno_df.sort_values('__parsed_date')
                                labels = turno_df['__parsed_date'].dt.strftime('%d/%m/%Y').tolist()
                                exec_values = turno_df['execucao_count'].astype(int).tolist()
                                treinado_values = turno_df['treinado_count'].astype(int).tolist()
                                merge_turno_charts[key]['labels'] = labels
                                merge_turno_charts[key]['values'] = exec_values
                                merge_turno_charts[key]['datasets'] = [
                                    {
                                        'key': 'execucao',
                                        'label': 'Execução por Voz (Sim)',
                                        'values': exec_values,
                                        'color': '#2563eb'
                                    },
                                    {
                                        'key': 'treinado',
                                        'label': 'Treinado (Sim)',
                                        'values': treinado_values,
                                        'color': '#16a34a'
                                    }
                                ]
                                merge_turno_charts[key]['totals'] = {
                                    'execucao': int(sum(exec_values)),
                                    'treinado': int(sum(treinado_values))
                                }
                            else:
                                merge_turno_charts[key]['labels'] = []
                                merge_turno_charts[key]['values'] = []
                                merge_turno_charts[key]['datasets'] = []
                                merge_turno_charts[key]['totals'] = {'execucao': 0, 'treinado': 0}

                preserved_args = build_query_args(overrides={'tab': 'input', 'input_filter': 'separacao'})
                preserved_args.pop('input_page', None)
                window = 2
                start_page = max(1, input_table_page - window)
                end_page = min(input_table_pages, input_table_page + window)

                page_links = []
                for p in range(start_page, end_page + 1):
                    args = {**preserved_args, 'input_page': p}
                    page_links.append({
                        'page': p,
                        'url': url_for('main.painel_grafico', **args),
                        'active': p == input_table_page
                    })

                input_table_pagination = {
                    'page': input_table_page,
                    'pages': input_table_pages,
                    'total': input_table_total,
                    'has_prev': input_table_page > 1,
                    'has_next': input_table_page < input_table_pages,
                    'prev_url': url_for(
                        'main.painel_grafico',
                        **{**preserved_args, 'input_page': input_table_page - 1}
                    ) if input_table_page > 1 else None,
                    'next_url': url_for(
                        'main.painel_grafico',
                        **{**preserved_args, 'input_page': input_table_page + 1}
                    ) if input_table_page < input_table_pages else None,
                    'page_links': page_links
                }
        except Exception as e:
            current_app.logger.exception('Falha ao preparar dados do Input*Dados: %s', e)

    input_column_meta = []
    input_filter_keys = set()
    for col in input_column_definitions:
        input_filter_keys.add(f"input_filter_{col['param']}")
        input_column_meta.append({
            'name': col['name'],
            'param': col['param'],
            'icon': col['icon'],
            'placeholder': col.get('placeholder', ''),
            'filter_value': col.get('filter_value', ''),
            'is_sorted': input_sort == col['param'],
            'sort_direction': input_order if input_sort == col['param'] else None,
        })

    input_form_args = build_query_args(
        skip_keys={'input_page', 'input_sort', 'input_order'} | input_filter_keys,
        overrides={'tab': 'input', 'input_filter': 'separacao'}
    )
    input_export_args = build_query_args(
        skip_keys={'input_page'},
        overrides={'tab': 'input', 'input_filter': 'separacao'}
    )

    hc_merged_rows = []
    hc_merged_columns = []
    hc_preview_info = None
    hc_training_chart = None
    hc_table_total = 0
    hc_table_page_size = 10
    hc_table_page = request.args.get('hc_page', default=1, type=int) or 1
    hc_table_page = max(1, hc_table_page)
    hc_table_pages = 0
    hc_table_range_start = 0
    hc_table_range_end = 0
    hc_table_pagination = None
    hc_sort = (request.args.get('hc_sort') or '').strip()
    hc_order = (request.args.get('hc_order') or 'asc').lower()
    if hc_order not in {'asc', 'desc'}:
        hc_order = 'asc'
    hc_column_meta = []
    hc_slug_to_column = {}

    try:
        source_hc = last_planilha_hc
    except NameError:
        source_hc = None

    if source_hc is not None:
        try:
            bind = db.session.get_bind()
            stmt = select(
                Colaborador.matricula.label("Matrícula"),
                Colaborador.nome.label("Nome"),
                Colaborador.tipo.label("Tipo"),
                Colaborador.setor.label("Setor"),
                Colaborador.area.label("Área"),
                Colaborador.turno.label("Turno"),
                Colaborador.supervisor.label("Supervisor"),
                Colaborador.integracao.label("Integração"),
                Colaborador.data.label("Data"),
            )
            df_db = pd.read_sql(stmt, bind)
            df_db['Matrícula'] = df_db['Matrícula'].apply(normalize_matricula)
            df_db = df_db[df_db['Matrícula'].notna()].copy()
            try:
                df_db['Matrícula'] = df_db['Matrícula'].astype(int)
            except Exception:
                pass

            merged_hc = pd.merge(df_db, source_hc, on='Matrícula', how='left')

            execucao_lookup = build_execucao_por_voz_lookup(source_df)
            if execucao_lookup is not None:
                merged_hc = pd.merge(
                    merged_hc,
                    execucao_lookup,
                    on='Matrícula',
                    how='left'
                )
            hc_plan_for_chart = source_hc[['Matrícula', 'Situação HC']].copy()
            if execucao_lookup is not None:
                hc_plan_for_chart = pd.merge(
                    hc_plan_for_chart,
                    execucao_lookup,
                    on='Matrícula',
                    how='left'
                )
            elif 'Execução por Voz' not in hc_plan_for_chart.columns:
                hc_plan_for_chart['Execução por Voz'] = ''
            if 'Situação HC' in merged_hc.columns:
                merged_hc['Situação HC'] = merged_hc['Situação HC'].apply(normalize_situacao_hc)
            if 'Situação HC' in hc_plan_for_chart.columns:
                hc_plan_for_chart['Situação HC'] = hc_plan_for_chart['Situação HC'].apply(normalize_situacao_hc)
            if {'Turno HC', 'Turno', 'Situação HC'}.issubset(merged_hc.columns):
                temporario_mask = merged_hc['Situação HC'] == 'Tempórario'
                if temporario_mask.any():
                    merged_hc.loc[temporario_mask, 'Turno HC'] = merged_hc.loc[temporario_mask, 'Turno']
            cargo_hc_column = 'Cargo HC' if 'Cargo HC' in merged_hc.columns else None
            if cargo_hc_column:
                with_hc = int(merged_hc[cargo_hc_column].notna().sum())
                without_hc = int(merged_hc[cargo_hc_column].isna().sum())
            else:
                with_hc = 0
                without_hc = len(merged_hc)

            hc_preview_info = {
                'total': len(merged_hc),
                'with_hc': with_hc,
                'without_hc': without_hc,
            }

            execucao_column = 'Execução por Voz' if 'Execução por Voz' in merged_hc.columns else None
            if execucao_column and {'Situação HC', 'Matrícula'}.issubset(merged_hc.columns):
                pivot_source = merged_hc[['Situação HC', execucao_column, 'Matrícula']].copy()

                def _clean_execucao(value):
                    if pd.isna(value):
                        return ''
                    text = str(value).strip()
                    lowered = text.lower()
                    if not text or lowered in {'nan', 'none', 'null', 'sem informação', 'sem informacao', 'sem dados'}:
                        return ''
                    return text

                def _normalize_execucao_category(value):
                    if not value:
                        return ''
                    text = str(value).strip()
                    if not text:
                        return ''
                    normalized = unicodedata.normalize('NFKD', text).encode('ASCII', 'ignore').decode('ASCII')
                    normalized = re.sub(r'\s+', ' ', normalized).strip().lower()
                    if not normalized:
                        return ''
                    normalized = normalized.replace('%', '')
                    if 'nao' in normalized:
                        return 'Não'
                    if 'sim' in normalized:
                        return 'Sim'
                    return ''

                pivot_source[execucao_column] = pivot_source[execucao_column].apply(_clean_execucao)
                pivot_source = pivot_source[pivot_source[execucao_column] != '']
                pivot_source[execucao_column] = pivot_source[execucao_column].apply(_normalize_execucao_category)
                pivot_source = pivot_source[pivot_source[execucao_column] != '']
                if not pivot_source.empty:
                    pivot_source['Situação HC'] = pivot_source['Situação HC'].fillna('Sem Situação').astype(str).str.strip()
                    pivot_source.loc[pivot_source['Situação HC'] == '', 'Situação HC'] = 'Sem Situação'

                    pivot_table = pd.pivot_table(
                        pivot_source,
                        index='Situação HC',
                        columns=execucao_column,
                        values='Matrícula',
                        aggfunc='count',
                        fill_value=0,
                    )

                    if not pivot_table.empty:
                        try:
                            pivot_table = pivot_table.astype(int)
                        except Exception:
                            pivot_table = pivot_table.applymap(lambda x: int(x) if pd.notna(x) else 0)

                        desired_execucao = ['Sim', 'Não']
                        pivot_table = pivot_table.loc[:, [col for col in pivot_table.columns if col in desired_execucao]]
                        for col in desired_execucao:
                            if col not in pivot_table.columns:
                                pivot_table[col] = 0
                        pivot_table = pivot_table[desired_execucao]
                        pivot_table = pivot_table.loc[:, (pivot_table != 0).any(axis=0)]

                        if not pivot_table.empty:
                            totals_by_situacao = pivot_table.sum(axis=1).sort_values(ascending=False)
                            pivot_table = pivot_table.loc[totals_by_situacao.index]
                            pivot_table = pivot_table.loc[:, [col for col in desired_execucao if col in pivot_table.columns]]

                            datasets = []
                            for column in pivot_table.columns:
                                column_label = str(column).strip() or 'Execução não informada'
                                values = [int(v) for v in pivot_table[column].astype(int).tolist()]
                                datasets.append({
                                    'label': column_label,
                                    'data': values,
                                    'total': int(sum(values)),
                                })

                            hc_training_chart = {
                                'situacao_labels': [str(idx) for idx in pivot_table.index.tolist()],
                                'datasets': datasets,
                                'totals': [int(v) for v in totals_by_situacao.loc[pivot_table.index].astype(int).tolist()],
                                'execucao_labels': [str(c) for c in pivot_table.columns.tolist()],
                                'overall_total': int(pivot_table.values.sum()),
                            }


            slug_counts = {}
            hc_filters = {}
            for column in merged_hc.columns:
                base_slug = slugify_column(column)
                if base_slug in slug_counts:
                    slug_counts[base_slug] += 1
                    slug = f"{base_slug}_{slug_counts[base_slug]}"
                else:
                    slug_counts[base_slug] = 1
                    slug = base_slug
                hc_slug_to_column[slug] = column
                value = (request.args.get(f"hc_filter_{slug}") or '').strip()
                if value:
                    hc_filters[column] = value
                hc_column_meta.append({
                    'name': str(column),
                    'slug': slug,
                    'filter_value': value,
                })

            filtered_hc = merged_hc.copy()
            for column, value in hc_filters.items():
                filter_series = filtered_hc[column].astype(str).fillna('')
                filtered_hc = filtered_hc[filter_series.str.contains(value, case=False, na=False)]

            if hc_sort in hc_slug_to_column:
                sort_column = hc_slug_to_column[hc_sort]
                ascending = hc_order == 'asc'
                filtered_hc = sort_dataframe(filtered_hc, sort_column, ascending=ascending)
            else:
                hc_sort = ''
                hc_order = 'asc'

            display_df = filtered_hc.fillna('')
            hc_table_total = len(display_df)
            if hc_table_total > 0:
                hc_table_pages = max(1, math.ceil(hc_table_total / hc_table_page_size))
                if hc_table_page > hc_table_pages:
                    hc_table_page = hc_table_pages
                start = (hc_table_page - 1) * hc_table_page_size
                end = start + hc_table_page_size
                page_df = display_df.iloc[start:end].copy()
                hc_table_range_start = start + 1
                hc_table_range_end = min(end, hc_table_total)

                hc_merged_columns = [str(c) for c in list(display_df.columns)]
                try:
                    hc_merged_rows = page_df.astype(str).values.tolist()
                except Exception:
                    hc_merged_rows = page_df.values.tolist()

                preserved_args = build_query_args(overrides={'tab': 'input', 'input_filter': 'hc'})
                preserved_args.pop('hc_page', None)

                window = 2
                start_page = max(1, hc_table_page - window)
                end_page = min(hc_table_pages, hc_table_page + window)

                page_links = []
                for p in range(start_page, end_page + 1):
                    args = {**preserved_args, 'hc_page': p}
                    page_links.append({
                        'page': p,
                        'url': url_for('main.painel_grafico', **args),
                        'active': p == hc_table_page
                    })

                hc_table_pagination = {
                    'page': hc_table_page,
                    'pages': hc_table_pages,
                    'total': hc_table_total,
                    'has_prev': hc_table_page > 1,
                    'has_next': hc_table_page < hc_table_pages,
                    'prev_url': url_for('main.painel_grafico', **{**preserved_args, 'hc_page': hc_table_page - 1}) if hc_table_page > 1 else None,
                    'next_url': url_for('main.painel_grafico', **{**preserved_args, 'hc_page': hc_table_page + 1}) if hc_table_page < hc_table_pages else None,
                    'page_links': page_links,
                }
            else:
                hc_merged_columns = [str(c) for c in list(display_df.columns)]
        except Exception as e:
            current_app.logger.exception('Falha ao gerar merge HC: %s', e)

    hc_filter_keys = set()
    for meta in hc_column_meta:
        slug = meta['slug']
        hc_filter_keys.add(f"hc_filter_{slug}")
        meta['is_sorted'] = hc_sort == slug
        meta['sort_direction'] = hc_order if meta['is_sorted'] else None

    hc_form_args = build_query_args(
        skip_keys={'hc_page', 'hc_sort', 'hc_order'} | hc_filter_keys,
        overrides={'tab': 'input', 'input_filter': 'hc'}
    )
    hc_export_args = build_query_args(
        skip_keys={'hc_page'},
        overrides={'tab': 'input', 'input_filter': 'hc'}
    )

    return render_template(
        'painel_grafico.html',
        total_colaboradores=total_colaboradores,
        min_data=min_data,
        max_data=max_data,
        min_all_str=min_all_str,
        max_all_str=max_all_str,
        min_data_str=min_data_str,
        max_data_str=max_data_str,
        today_str=today_str,
        setor_labels=setor_labels,
        setor_series=setor_series,
        tipo_labels=tipo_labels,
        tipo_series=tipo_series,
        turno_labels=turno_labels,
        turno_series=turno_series,
        stacked_categories=stacked_categories,
        stacked_series=stacked_series,
        timeline_data=timeline_data,
        available_turnos=available_turnos,
        selected_turno=selected_turno,
        available_setores=available_setores,
        available_tipos=available_tipos,
        available_supervisores=available_supervisores,
        selected_setor=selected_setor,
        selected_tipo=selected_tipo,
        selected_supervisor=selected_supervisor,
        input_table_columns=input_table_columns,
        input_table_rows=input_table_rows,
        input_table_total=input_table_total,
        input_table_page=input_table_page,
        input_table_pages=input_table_pages,
        input_table_page_size=input_table_page_size,
        input_table_has_data=input_table_has_data,
        input_table_pagination=input_table_pagination,
        input_table_range_start=input_table_range_start,
        input_table_range_end=input_table_range_end,
        merge_colab_percent=merge_colab_percent,
        merge_turno_charts=merge_turno_charts,
        input_column_meta=input_column_meta,
        input_sort=input_sort,
        input_order=input_order,
        input_form_args=input_form_args,
        input_export_args=input_export_args,
        hc_merged_columns=hc_merged_columns,
        hc_merged_rows=hc_merged_rows,
        hc_preview_info=hc_preview_info,
        hc_table_total=hc_table_total,
        hc_table_page=hc_table_page,
        hc_table_pages=hc_table_pages,
        hc_table_page_size=hc_table_page_size,
        hc_table_pagination=hc_table_pagination,
        hc_table_range_start=hc_table_range_start,
        hc_table_range_end=hc_table_range_end,
        hc_column_meta=hc_column_meta,
        hc_sort=hc_sort,
        hc_order=hc_order,
        hc_form_args=hc_form_args,
        hc_export_args=hc_export_args,
        hc_training_chart=hc_training_chart,
    )


@bp.route('/painel-grafico/export/separacao', methods=['GET'])
def export_input_separacao():
    try:
        source_df = last_planilha
    except NameError:
        source_df = None

    if source_df is None:
        flash('Nenhuma planilha de separação carregada para exportação.', 'warning')
        return redirect(url_for('main.painel_grafico', tab='input', input_filter='separacao'))

    hc_turno_lookup = {}
    try:
        lookup_source_hc = last_planilha_hc
    except NameError:
        lookup_source_hc = None
    if lookup_source_hc is not None:
        try:
            if {'Matrícula', 'Turno HC'}.issubset(lookup_source_hc.columns):
                turno_lookup_df = lookup_source_hc[['Matrícula', 'Turno HC']].copy()
                turno_lookup_df['Matrícula'] = turno_lookup_df['Matrícula'].apply(normalize_matricula)
                turno_lookup_df = turno_lookup_df[turno_lookup_df['Matrícula'].notna()]
                turno_lookup_df['Turno HC'] = turno_lookup_df['Turno HC'].fillna('').astype(str).str.strip()
                turno_lookup_df = turno_lookup_df.drop_duplicates(subset=['Matrícula'], keep='first')
                hc_turno_lookup = turno_lookup_df.set_index('Matrícula')['Turno HC'].to_dict()
        except Exception as err:
            current_app.logger.warning('Falha ao construir lookup de Turno HC para exportação: %s', err)
            hc_turno_lookup = {}

    definitions = get_input_column_definitions()
    columns = [col['name'] for col in definitions]

    try:
        export_df = source_df.copy()
    except Exception as err:
        current_app.logger.exception('Falha ao copiar planilha de separação para exportação: %s', err)
        flash(f'Falha ao preparar dados para exportação: {err}', 'danger')
        return redirect(url_for('main.painel_grafico', tab='input', input_filter='separacao'))

    for column in columns:
        if column not in export_df.columns:
            default_value = 'Não' if column == 'Treinado' else ''
            export_df[column] = default_value

    export_df = export_df[columns].copy()

    if hc_turno_lookup and {'Funcionário', 'Turno HC'}.issubset(export_df.columns):
        try:
            normalized_matriculas = export_df['Funcionário'].apply(normalize_matricula)
            mapped_turnos = normalized_matriculas.map(hc_turno_lookup).fillna('')
            existing_turnos = export_df['Turno HC'].fillna('').astype(str)
            export_df['Turno HC'] = existing_turnos.where(existing_turnos.str.strip() != '', mapped_turnos).fillna('')
        except Exception as err:
            current_app.logger.warning('Falha ao combinar Turno HC na exportação de Input*Dados: %s', err)

    filters = {}
    for definition in definitions:
        filter_value = (request.args.get(f"input_filter_{definition['param']}") or '').strip()
        if filter_value:
            filters[definition['name']] = filter_value

    for column, value in filters.items():
        try:
            series = export_df[column].astype(str).fillna('')
            export_df = export_df[series.str.contains(value, case=False, na=False)]
        except Exception as err:
            current_app.logger.warning('Falha ao aplicar filtro "%s" na exportação: %s', column, err)

    input_sort = (request.args.get('input_sort') or '').strip()
    input_order = (request.args.get('input_order') or 'asc').lower()
    valid_sorts = {col['param'] for col in definitions}
    if input_sort in valid_sorts:
        sort_column = next((col['name'] for col in definitions if col['param'] == input_sort), None)
        if sort_column:
            ascending = input_order != 'desc'
            try:
                export_df = sort_dataframe(export_df, sort_column, ascending=ascending)
            except Exception as err:
                current_app.logger.warning('Falha ao ordenar exportação por %s: %s', sort_column, err)

    export_df = export_df.fillna('')

    try:
        return dataframe_to_excel_response(export_df, filename_prefix='input_dados', sheet_name='Separacao')
    except RuntimeError as err:
        flash(str(err), 'danger')
        return redirect(url_for('main.painel_grafico', tab='input', input_filter='separacao'))


@bp.route('/painel-grafico/export/hc', methods=['GET'])
def export_input_hc():
    try:
        source_hc = last_planilha_hc
    except NameError:
        source_hc = None

    if source_hc is None:
        flash('Nenhuma planilha HC carregada para exportação.', 'warning')
        return redirect(url_for('main.painel_grafico', tab='input', input_filter='hc'))

    try:
        bind = db.session.get_bind()
        stmt = select(
            Colaborador.matricula.label("Matrícula"),
            Colaborador.nome.label("Nome"),
            Colaborador.tipo.label("Tipo"),
            Colaborador.setor.label("Setor"),
            Colaborador.area.label("Área"),
            Colaborador.turno.label("Turno"),
            Colaborador.supervisor.label("Supervisor"),
            Colaborador.integracao.label("Integração"),
            Colaborador.data.label("Data"),
        )
        df_db = pd.read_sql(stmt, bind)
    except Exception as err:
        current_app.logger.exception('Falha ao carregar dados do banco para exportação HC: %s', err)
        flash(f'Falha ao carregar dados do banco para exportação: {err}', 'danger')
        return redirect(url_for('main.painel_grafico', tab='input', input_filter='hc'))

    df_db['Matrícula'] = df_db['Matrícula'].apply(normalize_matricula)
    df_db = df_db[df_db['Matrícula'].notna()].copy()
    try:
        df_db['Matrícula'] = df_db['Matrícula'].astype(int)
    except Exception:
        pass

    try:
        merged_hc = pd.merge(df_db, source_hc, on='Matrícula', how='left')
    except Exception as err:
        current_app.logger.exception('Falha ao gerar merge HC para exportação: %s', err)
        flash(f'Falha ao mesclar dados do banco com HC: {err}', 'danger')
        return redirect(url_for('main.painel_grafico', tab='input', input_filter='hc'))

    try:
        source_df = last_planilha
    except NameError:
        source_df = None

    execucao_lookup = build_execucao_por_voz_lookup(source_df)
    if execucao_lookup is not None:
        merged_hc = pd.merge(merged_hc, execucao_lookup, on='Matrícula', how='left')

    if 'Situação HC' in merged_hc.columns:
        merged_hc['Situação HC'] = merged_hc['Situação HC'].apply(normalize_situacao_hc)
    if {'Turno HC', 'Turno', 'Situação HC'}.issubset(merged_hc.columns):
        temporario_mask = merged_hc['Situação HC'] == 'Tempórário'
        if temporario_mask.any():
            merged_hc.loc[temporario_mask, 'Turno HC'] = merged_hc.loc[temporario_mask, 'Turno']

    slug_counts = {}
    slug_to_column = {}
    filters = {}
    for column in merged_hc.columns:
        base_slug = slugify_column(column)
        count = slug_counts.get(base_slug, 0)
        if count:
            slug = f"{base_slug}_{count + 1}"
        else:
            slug = base_slug
        slug_counts[base_slug] = count + 1
        slug_to_column[slug] = column
        value = (request.args.get(f"hc_filter_{slug}") or '').strip()
        if value:
            filters[column] = value

    export_df = merged_hc.copy()
    for column, value in filters.items():
        try:
            series = export_df[column].astype(str).fillna('')
            export_df = export_df[series.str.contains(value, case=False, na=False)]
        except Exception as err:
            current_app.logger.warning('Falha ao aplicar filtro "%s" na exportação HC: %s', column, err)

    hc_sort = (request.args.get('hc_sort') or '').strip()
    hc_order = (request.args.get('hc_order') or 'asc').lower()
    if hc_sort in slug_to_column:
        sort_column = slug_to_column[hc_sort]
        ascending = hc_order != 'desc'
        try:
            export_df = sort_dataframe(export_df, sort_column, ascending=ascending)
        except Exception as err:
            current_app.logger.warning('Falha ao ordenar exportação HC por %s: %s', sort_column, err)

    export_df = export_df.fillna('')

    try:
        return dataframe_to_excel_response(export_df, filename_prefix='merge_hc', sheet_name='MergeHC')
    except RuntimeError as err:
        flash(str(err), 'danger')
        return redirect(url_for('main.painel_grafico', tab='input', input_filter='hc'))